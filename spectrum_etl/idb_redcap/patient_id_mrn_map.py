'''
Created on June 28, 2019

@author: pashaa@mskcc.org

Get relevant info from REDCap and then merge with IDB excel dump to create a new dump. Then add to REDCap.

'''


#!/usr/bin/env python
import pycurl, StringIO
from csv import reader

import pandas as pd
from openpyxl import load_workbook
from pandas import DataFrame

buf = StringIO.StringIO()
data = {
    #'token': '2A87E333079C6E02ABA3942E28E60391',
    'token': 'BB2C90E36A08E78EC47A3689C97EFBFF',
    'content': 'record',
    'format': 'csv',
    'type': 'flat',
    # 'records[0]': 'SPECTRUM-OV-001',
    # 'records[1]': 'SPECTRUM-OV-002',
    # 'records[2]': 'SPECTRUM-OV-003',
    # 'records[3]': 'SPECTRUM-OV-004',
    # 'records[4]': 'SPECTRUM-OV-005',
    #'fields[0]': 'patient_id',
    #'fields[1]': 'laboratory_test_guid',
    'forms[0]': 'laboratory_tests',
    'events[0]': 'patient_consent_arm_1',
    'rawOrLabel': 'raw',
    # 'rawOrLabelHeaders': 'label',
    # 'exportCheckboxLabel': 'false',
    # 'exportSurveyFields': 'false',
    # 'exportDataAccessGroups': 'false',
     'returnFormat': 'json'
}
ch = pycurl.Curl()
ch.setopt(ch.URL, 'https://redcap.mskcc.org/api/')
ch.setopt(ch.HTTPPOST, data.items())
ch.setopt(ch.WRITEFUNCTION, buf.write)
ch.perform()
ch.close()

#print(buf.getvalue())

#get counts:
table = reader(buf.getvalue().splitlines(), delimiter=',', quotechar='"')

header = ''
count = 0

# get_header row
for row in table:
    if header == '':
        header = row
        break
print("\nheader: " + str(header))

num_cols = len(header)

# initialize count map
count_map = {}
for name in header:
    count_map[name] = 0

df = DataFrame(table, columns=header)
pd.set_option('display.max_colwidth', -1)
pd.set_option('display.max_columns', None)

print("\nShape(rows, cols): "+str(df.shape))


for ii in range(1, df.shape[0]):
    for jj in range(0, df.shape[1]):
        # For printing the second column
        elem = df.iloc[ii, jj]
        if not "".__eq__(elem.strip()):
            count_map[header[jj]] += 1

print("\nField counts (non-empty strings): ")
for field_name in header:
    print("{0:40} = {1}".format(field_name, count_map[field_name]))



# for row in df.iterrows():
#     print(">>>>>>>"+str(row))
#     col_index = 0
#     for elem in row:
#         if not "".__eq__(elem.strip()):
#             # if elem is non-empty string
#             count_map[header_row[col_index]] += 1
#             # if header_row[col_index] == 'laboratory_test_subname':
#             #     print('laboratory_test_subname: '+str(elem))
#
#         col_index += 1
#
#
#
# print('>>>>>'+str(count_map))
# print(header_row)


'''
count: 21458
['patient_id', 'redcap_event_name', 'redcap_repeat_instrument', 'redcap_repeat_instance', 'laboratory_test_guid']
SPECTRUM-OV-024,patient_consent_arm_1,laboratory_tests,166,9115573189600070
SPECTRUM-OV-024,patient_consent_arm_1,laboratory_tests,10001,9115298527000070
SPECTRUM-OV-024,patient_consent_arm_1,laboratory_tests,10002,9115298527100070
SPECTRUM-OV-024,patient_consent_arm_1,laboratory_tests,10003,9115298527200070
SPECTRUM-OV-024,patient_consent_arm_1,laboratory_tests,10004,9115298534800070
'''


''' build mrn <-> id map'''
# redcap_reader = reader(buf.getvalue().splitlines(), delimiter=',', quotechar='"')
# id_to_mrn = {}  # patient id to mrn map
# mrn_to_id = {}
# is_header_row = True
# for row in redcap_reader:
#     if is_header_row:
#         is_header_row = False  # passed header row
#     else:
#         id_to_mrn[row[0]] = row[4]
#         mrn_to_id[row[4]] = row[0]
#
# print(id_to_mrn)
# print(mrn_to_id)

''' get idb result set '''

# wb = load_workbook("sample_data/Sample Data - SKI18134 (2019-07-05).xlsx")
#
# try:
#     treatments = wb['treatments']
#
#     rows = treatments.iter_rows(min_row=1, max_row=1)  # returns a generator of rows
#     first_row = next(rows)  # get the first row
#     header = [c.value for c in first_row]  # extract the values from the cells
#
#     df = DataFrame(treatments.values, columns=header)
#     df = df.drop([0])
#
# except NameError as ex:
#     raise ValueError('Error: worksheet not found %s' % (ex))
#
# ''' add cols for [patient_id, redcap_event_name, redcap_repeat_instrument, redcap_repeat_instance] '''
#
# df['redcap_event_name'] = 'treatment_arm_1'
# df['redcap_repeat_instrument'] = 'treatments'
#
# df.set_index('patient_mrn', inplace=True)
#
# temp_df1 = df.loc['38045928', :]
# temp_df1['redcap_repeat_instance'] = [ii for ii in range(1, temp_df1.shape[0] + 1)]
#
# temp_df2 = df.loc['38063538', :]
# temp_df2['redcap_repeat_instance'] = [ii for ii in range(1, temp_df2.shape[0] + 1)]
#
# frames = [temp_df1, temp_df2]
#
# print(temp_df1)
# result = pd.concat(frames)
# print(result)
#
# buf.close()


