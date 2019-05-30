'''
Created on May 23, 2019

@author: pashaa@mskcc.org
'''
import re
from typing import List

from openpyxl.styles import PatternFill, Protection
from openpyxl.worksheet.worksheet import Worksheet

MRN_CELL: str = 'A2'
ID_CELL: str = 'B2'
ID_PATTERN = r"(^SPECTRUM-OV-\d\d\d$)"
COLOR_PROCESSED = 'CFE7F7'  # a light shade of blue

class Patient(object):
    '''
    A patient object.

    Invariants: patient_mrn and patient_id are never changed in the worksheet.
    '''

    patient_sheet: Worksheet
    patient_mrn: int
    patient_id: str

    def __init__(self, patient_sheet: Worksheet):
        '''
        pre-condition:
        :param patient_sheet: a non None patient_sheet
        '''

        self.patient_sheet = patient_sheet
        self.validate()


    def validate(self) -> None:
        '''
        pre-condition: a non None patient sheet.
        post-condition: a valid patient object or exception
        '''
        if self.patient_sheet is None:
            print("Error: Unable to find patient sheet in workbook!")

        # patient_mrn
        try:
            self.patient_mrn = int(self.patient_sheet[MRN_CELL].value)
        except ValueError as ex:
            print('%s %s must be an integer in patients tab. %s' %
                  (MRN_CELL, self.patient_sheet[MRN_CELL], ex))

        # patient_id
        id_val = str(self.patient_sheet[ID_CELL].value)

        match = re.match(ID_PATTERN, id_val)

        if match is None:
            raise Exception('Invalid format for patient_id "%s" in patients tab. Expecting "%s"' % (id_val, ID_PATTERN))
        else:
            self.patient_id = match.groups()[0]


    def get_mrn(self) -> int:
        '''
        pre-condition: none
        post-condition: a valid patient_mrn returned
        '''
        return self.patient_mrn

    def get_id(self) -> str:
        '''
        pre-condition: none
        post-condition: a valid patient_id returned
        '''
        return self.patient_id

    def set_patient_as_processed(self):
        '''
        pre-condition: None
        Post-condition: cells are colored blue, locked and worksheet is protected.
        '''
        # color cells
        self.patient_sheet[MRN_CELL].fill = PatternFill("solid", fgColor=COLOR_PROCESSED)
        self.patient_sheet[ID_CELL].fill = PatternFill("solid", fgColor=COLOR_PROCESSED)

        # lock cells
        self.patient_sheet[MRN_CELL].protection = Protection(locked=True, hidden=False)
        self.patient_sheet[ID_CELL].protection = Protection(locked=True, hidden=False)

        ####  get and set worksheet protection state  ###
        self.patient_sheet.protection.sheet = True
        self.patient_sheet.protection.enable()





