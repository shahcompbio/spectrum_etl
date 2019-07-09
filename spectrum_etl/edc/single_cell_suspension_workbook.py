'''
Created on May 23, 2019

@author: pashaa@mskcc.org
'''
from openpyxl.styles import PatternFill

from openpyxl import load_workbook, styles


# DONE get and set cell color
# DONE get and set worksheet protection state
# DONE read patient tab, check for one new patient entry and store value
# TODO read surgeries tab, check for one new surgery type entry, store value and generate surgery id
# TODO read specimen tab, get one or more specimen sites and counts, store and generate speciment ids
# TODO read generate aliquot ids
# TODO add logs
# TODO look into applying additional data validation (https://openpyxl.readthedocs.io/en/stable/validation.html)
from spectrum_etl.edc.Surgeries import Surgeries
from spectrum_etl.edc.patient import Patient


class SingleCellSuspensionWorkbook(object):
    '''
    An object that represents a single cell suspension workbook for electronic data capture. It contains methods for
    performing a set of automated processes on the workbook.

    Invariants: Manual entries in the workbook are never modified.
                Workbook is always protected.
    '''

    patient: Patient
    surgeries: Surgeries

    def __init__(self, path_to_workbook):
        '''
        pre-condition:
        @:param path_to_workbook relative or absolute path to a single cell suspension workbook.
        '''
        if path_to_workbook is None:
            raise Exception("Error: no workbook path specified. Please specify a path!")

        wb = load_workbook(path_to_workbook)

        try:
            self.patient = Patient(wb['patients'])
            self.surgeries = Surgeries(wb['surgeries'])
        except NameError as ex:
            raise ValueError('Error: worksheet not found %s' % (ex))

        surgery_sheet = wb['surgeries']

        surgery_type_cell_range = surgery_sheet['A2':'A10']

        # surgery id is patient_id + surgery_type + autoinc_integer
        for cell_tup in surgery_type_cell_range:
            cell = cell_tup[0]
            if cell.value is not None:
                    print(">>>>>"+str(cell.row))
                    key = 'B'+str(cell.row)
                    surgery_id_cell = surgery_sheet[key]
                    surgery_id_cell.value = 'XXXXX'
            else:
                print(">>" + str(cell.row))


            #print(str(cell[0].value))

        #print(">>>>>" + str(surgery_type_cell_range))
        #print(">>>>>" + str(surgery_id_cell_range))

        # for cell_tup in surgery_type_cell_range:
        #     print(">>>>>" + str(cell_tup[0].value))
        #
        # print("\n\n")
        # for cell_tup in surgery_id_cell_range:
        #     print(">>>>>" + str(cell_tup[0].value))





        # df = pd.read_excel("tests/spectrum_etl/edc/single_cell_suspension.xlsx",
        #                    sheet_name=None,
        #                    index_col=0,
        #                     dtype = {'Name': str, 'Value': str}
        #                    )  # all sheets
        #
        # #self.patient = Patient(df['patient'])
        #
        # print(">>>>"+str(type(df)))
        # #print(">>>>"+str(self.patient.get_id()))
        #
        # # get surgery
        # # get specimens
        # # get aliquots
        #
        # # get surgery type
        # # populate surgery ids
        #
        #
        # # write to new excel sheet
        # writer = pd.ExcelWriter('single_cell_suspension_v1.xlsx', engine='xlsxwriter')
        #
        # # Convert the dataframe to an XlsxWriter Excel object.
        # df.to_excel(writer, sheet_name=None) # all sheets
        #
        # # Close the Pandas Excel writer and output the Excel file.
        # writer.save()
        #
        # # test with pre-existing data



