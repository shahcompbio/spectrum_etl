#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
Created on May 29, 2019

@author: pashaa@mskcc.org
'''
from openpyxl import load_workbook
import pytest
from openpyxl.styles import Protection

from spectrum_etl.edc.patient import Patient
from spectrum_etl.edc.patient import MRN_CELL
from spectrum_etl.edc.patient import ID_CELL
from spectrum_etl.edc.constants import COLOR_PROCESSED


class TestPatient(object):
    """Tests for `edc.patient` module."""
    @pytest.fixture(autouse=True)
    def setup_method(self, tmpdir):
        """ setup any state tied to the execution of the given method in a
        class.  setup_method is invoked for every test method of a class.
        """
        self.tmpdir = tmpdir



    def teardown_method(self, tmpdir):
        """ teardown any state that was previously setup with a setup_method
        call.
        """
        self.tmpdir = None

    def test_valid_patient(self):
        wb = load_workbook('tests/spectrum_etl/edc/data/patient_valid.xlsx')
        patient_sheet = wb['patients']

        patient = Patient(patient_sheet)
        assert patient.get_mrn() == 12345
        assert patient.get_id() == 'SPECTRUM-OV-001'

    def test_invalid_patient_mrn(self):
        wb = load_workbook('tests/spectrum_etl/edc/data/patient_invalid_mrn.xlsx')
        patient_sheet = wb['patients']

        with pytest.raises(ValueError, match=r".*must be an integer in patients tab.*"):
            Patient(patient_sheet)

    def test_invalid_patient_id(self):
        wb = load_workbook('tests/spectrum_etl/edc/data/patient_invalid_id.xlsx')
        patient_sheet = wb['patients']

        with pytest.raises(ValueError, match=r".*invalid format for patient_id in patients tab.*"):
            Patient(patient_sheet)


    def test_set_patient_as_processed(self):
        wb = load_workbook('tests/spectrum_etl/edc/data/patient_as_processed.xlsx')
        patient_sheet = wb['patients']
        assert patient_sheet[MRN_CELL].fill.fgColor.value == '00000000'  # no color
        assert patient_sheet[ID_CELL].fill.fgColor.value == '00000000'  # no color
        assert patient_sheet[MRN_CELL].protection.locked is False  # unlocked
        assert patient_sheet[ID_CELL].protection.locked is False  # unlocked
        assert patient_sheet.protection.sheet is False  # unprotected


        patient = Patient(patient_sheet)
        patient.set_sheet_as_processed()

        assert patient_sheet[MRN_CELL].fill.fgColor.value == '00'+COLOR_PROCESSED  # processed color
        assert patient_sheet[ID_CELL].fill.fgColor.value == '00'+COLOR_PROCESSED  # processed color
        assert patient_sheet[MRN_CELL].protection.locked is True  # locked
        assert patient_sheet[ID_CELL].protection.locked is True  # locked
        assert patient_sheet.protection.sheet is True  # unprotected


        wb.save('tests/spectrum_etl/edc/data/patient_as_processed_output.xlsx')
